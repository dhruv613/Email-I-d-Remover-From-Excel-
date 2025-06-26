from openpyxl import load_workbook
import os

# List of emails to remove
emails_to_remove = [

]

# Main function
def remove_emails_from_excel(file_path):
    if not os.path.exists(file_path):
        print("❌ File not found.")
        return

    # Normalize emails
    emails_to_remove_clean = [email.strip().lower() for email in emails_to_remove]

    # Load Excel
    wb = load_workbook(file_path)
    ws = wb.active

    # Find 'Email' column
    email_col_index = None
    for cell in ws[1]:  # Header row
        if str(cell.value).strip().lower() == "email":
            email_col_index = cell.column
            break

    if not email_col_index:
        print("❌ 'Email' column not found.")
        return

    # Scan for matching emails
    found_emails = []
    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=email_col_index).value
        if cell_value and cell_value.strip().lower() in emails_to_remove_clean:
            found_emails.append(cell_value.strip())

    if not found_emails:
        print("⚠️ No matching emails found.")
        return

    print("🟡 Found the following email(s) in the file:")
    for email in found_emails:
        print(f"   ➤ {email}")

    input("🔔 Press Enter to remove these emails from the file...")

    try:
        rows_deleted = 0
        for row in range(ws.max_row, 1, -1):
            cell_value = ws.cell(row=row, column=email_col_index).value
            if cell_value and cell_value.strip().lower() in emails_to_remove_clean:
                ws.delete_rows(row)
                rows_deleted += 1
        wb.save(file_path)
    except PermissionError as e:
        print(f"❌ Error while deleting rows: {e}")
        return

    print(f"✅ Removed {rows_deleted} email(s) from '{file_path}':")
    for email in found_emails:
        print(f"   ✔️ {email}")

# 📁 Call the function
remove_emails_from_excel("15k 23-6 n3_backup.xlsx")
